import RPi.GPIO as GPIO
import smbus
import time
from math import sqrt
from datetime import datetime, date, timedelta
import openpyxl



# Seleccionar el bus I2C
bus = smbus.SMBus(1)

# Direcicón I2C del dispositivo
MMA7660FC_DEFAULT_ADDRESS           = 0x4C

# MMA7660FC Register Map
MMA7660FC_XOUT                      = 0x00 # Output Value X
MMA7660FC_YOUT                      = 0x01 # Output Value Y
MMA7660FC_ZOUT                      = 0x02 # Output Value Z
MMA7660FC_TILT                      = 0x03 # Tilt Status
MMA7660FC_SRST                      = 0x04 # Sampling Rate Status
MMA7660FC_SPCNT                     = 0x05 # Sleep Count
MMA7660FC_INTSU                     = 0x06 # Interrupt Status
MMA7660FC_MODE                      = 0x07 # Mode Register
MMA7660FC_SR                        = 0x08 # Sample Rate Register
MMA7660FC_PDET                      = 0x09 # Tap/Pulse Detection Register
MMA7660FC_PD                        = 0x0A # Tap/Pulse Debounce Count Register

# MMA7660FC Mode Register
MMA7660FC_MODE_STANDBY              = 0x00 # Standby Mode
MMA7660FC_MODE_TEST                 = 0x04 # Test Mode
MMA7660FC_MODE_ACTIVE               = 0x01 # Active Mode
MMA7660FC_AWE_EN                    = 0x08 # Auto-Wake Enabled
MMA7660FC_AWE_DS                    = 0x00 # Auto-Wake Disabled
MMA7660FC_ASE_EN                    = 0x10 # Auto-Sleep Enabled
MMA7660FC_ASE_DS                    = 0x00 # Auto-Sleep Disabled
MMA7660FC_SCPS_16                   = 0x20 # Prescaler is divide by 16
MMA7660FC_SCPS_1                    = 0x00 # Prescaler is divide by 1
MMA7660FC_IPP_OPEN                  = 0x00 # Interrupt output INT is open-drain
MMA7660FC_IPP_PUSH                  = 0x40 # Interrupt output INT is push-pull
MMA7660FC_IAH_LOW                   = 0x00 # Interrupt output INT is active low
MMA7660FC_IAH_HIGH                  = 0x80 # Interrupt output INT is active high

# MMA7660FC Sample Rate Register
MMA7660FC_AMSR_120                  = 0x00 # 120 Samples/Second Active and Auto-Sleep Mode
MMA7660FC_AMSR_64                   = 0x01 # 64 Samples/Second Active and Auto-Sleep Mode
MMA7660FC_AMSR_32                   = 0x02 # 32 Samples/Second Active and Auto-Sleep Mode
MMA7660FC_AMSR_16                   = 0x03 # 16 Samples/Second Active and Auto-Sleep Mode
MMA7660FC_AMSR_8                    = 0x04 # 8 Samples/Second Active and Auto-Sleep Mode
MMA7660FC_AMSR_4                    = 0x05 # 4 Samples/Second Active and Auto-Sleep Mode
MMA7660FC_AMSR_2                    = 0x06 # 2 Samples/Second Active and Auto-Sleep Mode
MMA7660FC_AMSR_1                    = 0x07 # 1 Samples/Second Active and Auto-Sleep Mode
MMA7660FC_AWSR_32                   = 0x00 # 32 Samples/Second Auto-Wake Mode
MMA7660FC_AWSR_16                   = 0x08 # 16 Samples/Second Auto-Wake Mode
MMA7660FC_AWSR_8                    = 0x10 # 8 Samples/Second Auto-Wake Mode
MMA7660FC_AWSR_1                    = 0x18 # 1 Samples/Second Auto-Wake Modu



class MMA7660FC():
    def __init__(self):
        self.mode_config()
        self.sample_rate_config()
        self.pasos = 0    
    #Función para configurar el sensor antes de proceder a las medidas
    def mode_config(self):
        """Select the mode control register of the accelerometer from the given provided values"""
        MODE_CONTROL = (MMA7660FC_MODE_ACTIVE | MMA7660FC_AWE_DS | MMA7660FC_ASE_DS | MMA7660FC_SCPS_1 | MMA7660FC_IAH_LOW)
        bus.write_byte_data(MMA7660FC_DEFAULT_ADDRESS, MMA7660FC_MODE, MODE_CONTROL)
        
    #Función para configurar el registro de interrupciones
    def interrupt_config(self):
        bus.write_byte_data(MMA7660FC_DEFAULT_ADDRESS, MMA7660FC_INTSU, 0xE7)
    
    #Función para configurar el número de muestras por segundo
    def sample_rate_config(self):
        SAMPLE_RATE = (MMA7660FC_AMSR_2)
        bus.write_byte_data(MMA7660FC_DEFAULT_ADDRESS, MMA7660FC_SR, SAMPLE_RATE)
    

    #Función para leer los valores de x,y,z del sensor
    #Parámetros de la función bus.read_i2c_block_data
        #  direcicón I2C del dispositivo (MMA7660FC_DEFAULT_ADDRESS)
        #  donde se comienza a leer (MMA7660FC_XOUT)
        #  cuantos bytes se quiren leer (3)
    def read_accl(self):
        data = bus.read_i2c_block_data(MMA7660FC_DEFAULT_ADDRESS, MMA7660FC_XOUT, 3)
        
        # Se conveirten los datos a 6 bits, porque la trama leída completa son 8 bits, y los dos primeros
        # bits más significativos son de configuración. Los bits de la aceleracion son los bits 6-0
        xAccl = data[0] & 0x3F
        if xAccl > 31 :
            xAccl -= 64
        
        yAccl = data[1] & 0x3F
        if yAccl > 31 :
            yAccl -= 64
        
        zAccl = data[2] & 0x3F
        if zAccl > 31 :
            zAccl -= 64
        
        return {'x' : xAccl, 'y' : yAccl, 'z' : zAccl}



  #Función para leer los bits correspondientes a la orentación del sensor
    #Parámetros de la función bus.read_i2c_block_data
        #  direcicón I2C del dispositivo (MMA7660FC_DEFAULT_ADDRESS)
        #  donde se comienza a leer (MMA7660FC_TILT) -- que es 0x03
        #  cuantos bytes se quiren leer (1)
    def read_orientation(self):
        """Leer datos desde la posicion 0x03, 1 byte """
        data = bus.read_i2c_block_data(MMA7660FC_DEFAULT_ADDRESS, MMA7660FC_TILT, 1)
        
        PoLa = data[0] & 0x1C
        
        print(PoLa)
        if PoLa == 0x00:
            orientation = 0 #Desconocido
        elif PoLa == 0x04:
            print("Izquierda")
            orientation = 1  #Izquierda: Dispositivo esta en modo paisaje hacia la izquierda
        elif PoLa == 0x08:
            print("Derecha")
            orientation = 2  #Izquierda: Dispositivo esta en modo paisaje hacia la derecha
        elif PoLa == 0x14:
            print("Abajo")
            orientation = 3 # Abajo: Dispositivo está en posición vertical invertida
        elif PoLa == 0x18:
            print("Arriba")
            orientation = 4 # Arriba: Dispositivo está en posición vertical normal
        else:
            orientation = 0x00
        return orientation    


def setup():
     # ------------INICIALIZACIÓN Y CONFIGURACIÓN DEL SENSOR DE LUZ ----------------
    GPIO.cleanup()
    GPIO.setmode(GPIO.BCM)
    GPIO.setup(4,GPIO.IN)
     # ------------INICIALIZACIÓN Y CONFIGURACIÓN DEL LED VERDE ----------------
    GPIO.setup(18,GPIO.OUT, initial=GPIO.HIGH)
    GPIO.output(18,1)
     # ------------INICIALIZACIÓN Y CONFIGURACIÓN DEL LED ROJO ----------------
    GPIO.setup(23,GPIO.OUT, initial=GPIO.HIGH)
    GPIO.output(23,0)
    

def LED_V_ON():
    GPIO.output(18,GPIO.HIGH)

def LED_V_OFF():
    GPIO.output(18,GPIO.LOW)

def LED_R_ON():
    GPIO.output(23,GPIO.HIGH)
    time.sleep(0.1)
    GPIO.output(23,GPIO.LOW)
def LED_R_OFF():
    GPIO.output(23,GPIO.LOW)

n_filas_maximo = 50
n_fila = 1

def write_excel(fila, velocidad_x, velocidad_y, velocidad_z, velocidad, sensor):
    
    ruta = '/home/pi/Desktop/excelBD.xlsx'

    wb = openpyxl.load_workbook(ruta)

    sheet = wb.active

    # Añadir datos
    sheet.cell(row=fila, column=1).value = datetime.now().strftime('%Y/%m/%d %H:%M:%S')

    sheet.cell(row=fila, column=2).value = velocidad_x
    sheet.cell(row=fila, column=3).value = velocidad_y
    sheet.cell(row=fila, column=4).value = velocidad_z
    sheet.cell(row=fila, column=5).value = velocidad
    sheet.cell(row=fila, column=6).value = sensor
    
    
    print(n_fila)
    wb.save(ruta)


# ------------INICIALIZACIÓN Y CONFIGURACIÓN DEL ACELEROMETRO ----------------
mma7660fc = MMA7660FC()
mma7660fc.mode_config()
time.sleep(1)
mma7660fc.sample_rate_config()
time.sleep(1)
mma7660fc.interrupt_config()
time.sleep(1)

setup()

while True:

    time.sleep(1)
    accl = mma7660fc.read_accl()
    sensor = GPIO.input(4)


    #Tranformamos la lectura de mma7660fc.read_accl() a velocidad en ms^2

    velocidad_x = (accl['x']/21.33)/9.8
    velocidad_y = (accl['y']/21.33)/9.8
    velocidad_z = (accl['z']/21.33)/9.8

    velocidad = sqrt(velocidad_x**2 + velocidad_y**2 + velocidad_z**2)

    

    #Asumimos que si el sensor de luz devuelve 1 no detecta objeto 15 metros
    if sensor == 1:
        print("******************")
        print("No hay obstaculo")
        print("Vehiculo en marcha")
        print ("Velocidad: %f"%((velocidad/21.33)/9.8)+ " ms^2")
        print()
        LED_V_ON()
        LED_R_OFF()
        time.sleep(1)
    else:
        accl['y'] = 0
        print("******************")
        print("¡¡¡¡Obstaculo identificado!!!!")
        print ("Velocidad: " + str(velocidad_y)+" ms^2")
        print()
        LED_V_OFF()
        LED_R_ON()
        time.sleep(1)

    write_excel(n_fila,velocidad_x, velocidad_y, velocidad_z, velocidad, sensor)
    #Para sobreescribir en las 50 primeras filas
    n_fila = (n_fila + 1)%50